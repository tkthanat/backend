# app/main.py
import cv2
from fastapi import FastAPI, UploadFile, File, Form, WebSocket, WebSocketDisconnect, Depends, HTTPException, Response, \
    Body
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.sql import func
from sqlalchemy import or_
import os, io, csv, asyncio, base64, time, uuid, shutil
from datetime import datetime, date, timedelta, time as dt_time
from collections import defaultdict

import pandas as pd

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage

from .db_models import get_db, UserFace, User, AttendanceLog, Subject, UserType
from .camera_handler import CameraManager, discover_local_devices
from .ai_engine import refresh_facebank_from_db, load_facebank

# ✨ [เพิ่ม] 1. Import auth.py ที่เราสร้าง
from . import auth

from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

app = FastAPI(title="Offline Attendance (Minimal)")

COOLDOWN_SECONDS = 30

# --- 1. CORS Middleware ---
origins = ["http://localhost:3000", ]
app.add_middleware(
    CORSMiddleware, allow_origins=origins, allow_credentials=True, allow_methods=["*"], allow_headers=["*"],
)

# --- 2. Mount Static Directories ---
MEDIA_ROOT = os.getenv("MEDIA_ROOT", "./data/faces/train")
SNAPSHOTS_DIR = "media/snapshot"
os.makedirs(MEDIA_ROOT, exist_ok=True)
os.makedirs(SNAPSHOTS_DIR, exist_ok=True)
app.mount("/static", StaticFiles(directory="data"), name="static")
app.mount("/media/snapshot", StaticFiles(directory=SNAPSHOTS_DIR), name="static_snapshots")

# --- 3. Camera Manager Setup (เหมือนเดิม) ---
print("Discovering local devices for initial setup...")
discovered_devices = discover_local_devices(test_frame=False)
available_sources = [d['src'] for d in discovered_devices if d.get('opened', False)]
print(f"Available camera sources found: {available_sources}")
CAMERA_SOURCES = {}
if len(available_sources) > 0:
    CAMERA_SOURCES['entrance'] = available_sources[0]
else:
    CAMERA_SOURCES['entrance'] = "0"
if len(available_sources) > 1:
    CAMERA_SOURCES['exit'] = available_sources[1]
else:
    CAMERA_SOURCES['exit'] = CAMERA_SOURCES['entrance']
print(f"Assigning camera sources: {CAMERA_SOURCES}")
cam_mgr = CameraManager(CAMERA_SOURCES, fps=30, width=640, height=480)


@app.on_event("startup")
async def _startup():
    cnt = load_facebank()
    print(f"[facebank] loaded users={cnt}")


# --- ✨ [ใหม่] 4. Internal Helper Function ---
# (ย้าย Logic มาจาก /train/refresh เพื่อให้ฟังก์ชันอื่นเรียกใช้ได้)
def _internal_train_refresh(db: Session):
    """Helper function to run the training logic."""
    print("Running internal train refresh...")
    rows = (
        db.query(UserFace.user_id, UserFace.file_path, User.name)
        .join(User, User.user_id == UserFace.user_id).all()
    )
    users, total = refresh_facebank_from_db(rows)
    cnt = load_facebank()
    print(f"Internal train refresh complete: {users} users, {total} images, {cnt} loaded.")
    return {"message": "facebank updated", "users": users, "images_used": total, "loaded": cnt}


# --- 5. Face Upload & Training Endpoints ---
@app.post("/faces/upload")
async def upload_faces(
    user_id: int = Form(...),
    images: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    saved, items = 0, []
    user_dir = os.path.join(MEDIA_ROOT, str(user_id))
    os.makedirs(user_dir, exist_ok=True)
    for f in images:
        file_ext = os.path.splitext(f.filename)[1]
        name = f"{uuid.uuid4()}{file_ext}"
        dest = os.path.join(user_dir, name)
        content = await f.read()
        with open(dest, "wb") as wf: wf.write(content)
        uf = UserFace(user_id=user_id, file_path=name)
        db.add(uf)
        items.append({"file": name})
        saved += 1
    db.commit()
    return {"saved": saved, "items": items}


@app.post("/train/refresh")
def train_refresh(
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    # ✨ [แก้ไข] เรียกใช้ Helper แทน
    return _internal_train_refresh(db)


# --- 6. Camera Control & Stream Endpoints ---
@app.get("/cameras")
def list_cameras(
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    return {"cams": cam_mgr.list()}


@app.post("/cameras/{cam_id}/open")
def open_camera(
    cam_id: str,
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    cam_mgr.open(cam_id);
    return {"message": f"camera '{cam_id}' opened"}


@app.post("/cameras/{cam_id}/close")
def close_camera(
    cam_id: str,
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    cam_mgr.close(cam_id);
    return {"message": f"camera '{cam_id}' closed"}


@app.get("/cameras/{cam_id}/snapshot", responses={200: {"content": {"image/jpeg": {}}}})
def camera_snapshot(
    cam_id: str,
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    jpg = cam_mgr.get_jpeg(cam_id);
    return Response(content=jpg, media_type="image/jpeg")


# --- (เว้นไว้: mjpeg, ws_camera, ws_ai_results ไม่ต้องป้องกัน) ---
@app.get("/cameras/{cam_id}/mjpeg")
def camera_mjpeg(cam_id: str):
    boundary = "frame"

    async def gen():
        if cam_id not in cam_mgr.sources:
            print(f"MJPEG: Camera ID {cam_id} not found in sources.")
            return
        try:
            cam_mgr.open(cam_id)
            print(f"Opening MJPEG stream for {cam_id} (Source: {cam_mgr.sources[cam_id].src})")
            while True:
                try:
                    jpg = cam_mgr.get_jpeg(cam_id)
                    if not jpg:
                        await asyncio.sleep(0.1)
                        continue
                    yield (
                            b"--" + boundary.encode() + b"\r\n" + b"Content-Type: image/jpeg\r\n" + b"Cache-Control: no-cache\r\n" + b"Pragma: no-cache\r\n" + b"Content-Length: " + str(
                        len(jpg)).encode() + b"\r\n\r\n" + jpg + b"\r\n")
                except Exception as e:
                    if not cam_mgr.sources[cam_id].is_open:
                        print(f"MJPEG {cam_id} stopping because camera is no longer open.")
                        break
                    await asyncio.sleep(0.1)
                await asyncio.sleep(cam_mgr.interval)
        except Exception as e:
            print(f"Could not open camera {cam_id} for MJPEG: {e}")
        finally:
            print(f"Closing MJPEG stream for {cam_id}")
            cam_mgr.close(cam_id)

    return StreamingResponse(gen(), media_type=f"multipart/x-mixed-replace; boundary={boundary}",
                             headers={"Cache-Control": "no-cache, no-store, must-revalidate",
                                      "Connection": "keep-alive"})


@app.websocket("/ws/cameras/{cam_id}")
async def ws_camera(ws: WebSocket, cam_id: str):
    await ws.accept()
    try:
        cam_mgr.open(cam_id)
    except Exception:
        pass
    try:
        while True:
            await asyncio.sleep(0.1)
            try:
                jpg = cam_mgr.get_jpeg(cam_id)
                b64 = base64.b64encode(jpg).decode("ascii")
                await ws.send_json({"type": "frame", "cam_id": cam_id, "data": b64})
            except Exception as e:
                await ws.send_json({"type": "error", "message": str(e)})
    except WebSocketDisconnect:
        pass


@app.websocket("/ws/ai_results/{cam_id}")
async def ws_ai_results(ws: WebSocket, cam_id: str):
    await ws.accept()
    if cam_id not in cam_mgr.sources: await ws.close(code=1008, reason="Camera not found"); return
    cam = cam_mgr.sources[cam_id]
    if not cam.is_open:
        try:
            print(f"WS AI opening camera {cam_id}...")
            cam_mgr.open(cam_id)
        except Exception as e:
            await ws.close(code=1011, reason=f"Could not open camera: {e}");
            return
    print(f"[WS AI {cam_id}] Client connected.")
    try:
        while True:
            if not cam.is_open:
                print(f"[WS AI {cam_id}] Camera is closed, disconnecting.")
                break
            results = cam.last_ai_result
            await ws.send_json({"cam_id": cam_id, "results": results, "ai_width": cam_mgr.ai_process_width,
                                "ai_height": cam_mgr.ai_process_height})
            await asyncio.sleep(0.1)
    except WebSocketDisconnect:
        print(f"[WS AI {cam_id}] Client disconnected.")
    except Exception as e:
        print(f"[WS AI {cam_id}] Error: {e}")
    finally:
        print(f"[WS AI {cam_id}] Connection closed.")
# --- (สิ้นสุดส่วนที่ไม่ต้องป้องกัน) ---


@app.get("/cameras/discover")
async def cameras_discover(
    max_index: int = 10,
    test_frame: bool = True,
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    print("Discovering local devices...")
    active_sources = [c.src for c in cam_mgr.sources.values() if c.is_open]
    print(f"Active sources (will skip test): {active_sources}")
    devs = discover_local_devices(max_index=max_index, test_frame=test_frame, exclude_srcs=active_sources)
    print(f"Discovery found: {devs}")
    return {"devices": devs}


@app.get("/cameras/config")
def get_camera_config(
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    return {"mapping": {k: v.src for k, v in cam_mgr.sources.items()}}


@app.post("/cameras/config")
def set_camera_config(
    mapping: dict = Body(..., example={"entrance": "0", "exit": "1"}),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    print(f"Reconfiguring cameras to: {mapping}")
    cam_mgr.reconfigure(mapping)
    return {"message": "camera mapping updated", "mapping": mapping}


# --- 7. Attendance API Endpoints ---
class ActiveSubjectPayload(BaseModel):
    subject_id: Optional[int] = None

@app.post("/attendance/set_active_subject")
def set_active_subject(
    payload: ActiveSubjectPayload,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    active_user_ids: Optional[set] = None
    roster_size = 0
    if payload.subject_id is not None:
        users_in_subject = db.query(User.user_id).filter(
            User.subject_id == payload.subject_id,
            User.is_deleted == 0
        ).all()
        active_user_ids = {user.user_id for user in users_in_subject}
        roster_size = len(active_user_ids)
        print(f"[Attendance] Setting active subject {payload.subject_id}. Roster size: {roster_size}")
    else:
        print("[Attendance] Setting active subject to ALL.")
        active_user_ids = None
    cam_mgr.set_active_roster(active_user_ids, payload.subject_id)
    return {"message": "Active subject updated", "active_subject_id": payload.subject_id, "roster_size": roster_size}


@app.post("/attendance/start")
def start_attendance(
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    print("Starting AI processing for all cameras...")
    for cam in cam_mgr.sources.values(): cam.is_ai_paused = False
    return {"message": "Attendance started"}


@app.post("/attendance/stop")
def stop_attendance(
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    print("Stopping AI processing for all cameras...")
    for cam in cam_mgr.sources.values(): cam.is_ai_paused = True
    return {"message": "Attendance stopped"}


@app.get("/attendance/poll", response_model=List[dict])
async def get_attendance_events(
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    events = cam_mgr.get_attendance_events()
    if not events: return []
    today = date.today()
    new_logs_for_frontend = []
    user_ids_to_check = {e["user_id"] for e in events if e.get("user_id")}
    if not user_ids_to_check: return []
    users_data = db.query(User.user_id, User.subject_id, User.student_code, User.name).filter(
        User.user_id.in_(user_ids_to_check)).all()
    user_info_map = {u.user_id: u for u in users_data}
    active_subject_id = cam_mgr.active_subject_id
    active_rule_time: Optional[dt_time] = None
    if active_subject_id:
        subject = db.query(Subject.class_start_time).filter(Subject.subject_id == active_subject_id).first()
        if subject and subject.class_start_time:
            active_rule_time = subject.class_start_time
    for event in events:
        user_id = event.get("user_id")
        if not user_id or user_id not in user_info_map:
            continue
        user_info = user_info_map[user_id]
        log_subject_id: Optional[int] = None
        if active_subject_id is not None:
            log_subject_id = active_subject_id
        else:
            log_subject_id = user_info.subject_id
        event_timestamp = datetime.fromtimestamp(event["timestamp"])
        last_log = db.query(AttendanceLog).filter(
            AttendanceLog.user_id == user_id
        ).order_by(AttendanceLog.timestamp.desc()).first()
        can_log = False
        time_since_last_log = timedelta(days=1)
        if last_log:
            time_since_last_log = event_timestamp - last_log.timestamp
        if not last_log:
            if event["action"].lower() == "enter":
                can_log = True
        elif event["action"].lower() == "enter":
            if last_log.action.lower() == "exit":
                if time_since_last_log >= timedelta(seconds=COOLDOWN_SECONDS):
                    can_log = True
            elif last_log.action.lower() == "enter":
                if time_since_last_log >= timedelta(seconds=COOLDOWN_SECONDS):
                    can_log = True
        elif event["action"].lower() == "exit":
            if last_log.action.lower() == "enter":
                can_log = True
            elif last_log.action.lower() == "exit":
                if time_since_last_log >= timedelta(seconds=COOLDOWN_SECONDS):
                    can_log = True
        if can_log:
            log_status: Optional[str] = None
            log_rule: Optional[dt_time] = None
            if event["action"].lower() == "enter":
                log_status = "Present"
                log_rule = active_rule_time
                if log_rule:
                    class_start_dt = datetime.combine(event_timestamp.date(), log_rule)
                    if event_timestamp > class_start_dt:
                        log_status = "Late"
            new_log_db = AttendanceLog(
                user_id=user_id,
                subject_id=log_subject_id,
                action=event["action"],
                timestamp=event_timestamp,
                confidence=event.get("confidence"),
                log_status=log_status,
                log_rule_start_time=log_rule
            )
            if event["action"].lower() == "enter" and event.get("frame") is not None:
                os.makedirs(SNAPSHOTS_DIR, exist_ok=True)
                snap_name = f"{user_id}_{int(time.time())}.jpg"
                snap_path_absolute = os.path.join(SNAPSHOTS_DIR, snap_name)
                snap_path_relative = os.path.join("media/snapshot", snap_name).replace("\\", "/")
                try:
                    cv2.imwrite(snap_path_absolute, event["frame"])
                    new_log_db.snapshot_path = snap_path_relative
                except Exception as e:
                    print(f"Error saving snapshot: {e}")
                    new_log_db.snapshot_path = None
            db.add(new_log_db)
            db.flush()
            new_logs_for_frontend.append({
                "log_id": new_log_db.log_id,
                "user_id": user_id,
                "user_name": user_info.name,
                "student_code": user_info.student_code or "N/A",
                "action": event["action"],
                "timestamp": event_timestamp.isoformat(),
                "confidence": event.get("confidence"),
                "subject_id": log_subject_id,
                "snapshot_path": getattr(new_log_db, "snapshot_path", None),
                "log_status": log_status,
            })
    if new_logs_for_frontend:
        db.commit()
    return new_logs_for_frontend


@app.get("/attendance/logs", response_model=List[dict])
async def get_attendance_logs(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    subject_id: Optional[int] = None,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    query = (
        db.query(
            AttendanceLog, User.name.label("user_name"),
            User.student_code.label("student_code"), Subject.subject_name.label("subject_name")
        )
        .outerjoin(User, AttendanceLog.user_id == User.user_id)
        .outerjoin(Subject, AttendanceLog.subject_id == Subject.subject_id)
        .order_by(AttendanceLog.timestamp.desc())
    )
    query = query.filter(User.is_deleted != 1)
    if start_date: query = query.filter(func.date(AttendanceLog.timestamp) >= start_date)
    if end_date: query = query.filter(func.date(AttendanceLog.timestamp) <= end_date)
    if subject_id is not None: query = query.filter(AttendanceLog.subject_id == subject_id)
    logs = query.all()
    results = []
    for log, user_name, student_code, subject_name in logs:
        results.append({
            "log_id": log.log_id, "user_id": log.user_id,
            "subject_id": log.subject_id, "action": log.action,
            "timestamp": log.timestamp.isoformat(), "confidence": log.confidence,
            "user_name": user_name or "N/A", "student_code": student_code or "N/A",
            "subject_name": subject_name or None,
            "snapshot_path": log.snapshot_path if hasattr(log, 'snapshot_path') else None,
            "log_status": log.log_status,
            "log_rule_start_time": log.log_rule_start_time,
        })
    return results


@app.get("/attendance/export")
async def export_attendance_logs(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    subject_id: Optional[int] = None,
    format: str = "csv",
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    query = (
        db.query(
            AttendanceLog.timestamp.label("Timestamp"),
            User.student_code.label("StudentCode"),
            User.name.label("Name"),
            Subject.subject_name.label("Subject"),
            Subject.section.label("Section"),
            AttendanceLog.action.label("Action"),
            AttendanceLog.log_status.label("Status"),
            AttendanceLog.log_rule_start_time.label("RuleTime"),
            AttendanceLog.confidence.label("Confidence"),
            AttendanceLog.snapshot_path.label("SnapshotPath")
        )
        .outerjoin(User, AttendanceLog.user_id == User.user_id)
        .outerjoin(Subject, AttendanceLog.subject_id == Subject.subject_id)
        .order_by(AttendanceLog.timestamp.asc())
    )
    query = query.filter(User.is_deleted == 0)
    query = query.filter(or_(Subject.is_deleted == None, Subject.is_deleted == 0))
    if start_date: query = query.filter(func.date(AttendanceLog.timestamp) >= start_date)
    if end_date: query = query.filter(func.date(AttendanceLog.timestamp) <= end_date)
    if subject_id is not None: query = query.filter(AttendanceLog.subject_id == subject_id)
    logs = query.all()
    output = io.BytesIO()
    filename = f"attendance_export_{start_date or 'all'}_to_{end_date or 'all'}"
    headers_excel = ["Timestamp", "StudentCode", "Name", "Subject", "Section", "Action", "Status", "RuleTime",
                     "Confidence", "Snapshot"]
    headers_csv = ["Timestamp", "StudentCode", "Name", "Subject", "Section", "Action", "Status", "RuleTime",
                   "Confidence", "SnapshotPath"]
    if format.lower() == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(headers_excel)
        ws.column_dimensions['A'].width = 22;
        ws.column_dimensions['B'].width = 15;
        ws.column_dimensions['C'].width = 25;
        ws.column_dimensions['D'].width = 20;
        ws.column_dimensions['E'].width = 10;
        ws.column_dimensions['F'].width = 10;
        ws.column_dimensions['G'].width = 10;
        ws.column_dimensions['H'].width = 10;
        ws.column_dimensions['I'].width = 12;
        ws.column_dimensions['J'].width = 20
        if not logs:
            ws.append(["No data found for the selected filters."] + [""] * 9)
        else:
            for idx, log in enumerate(logs):
                row_num = idx + 2
                timestamp_str = pd.to_datetime(log.Timestamp).tz_localize(None).strftime('%Y-%m-%d %H:%M:%S')
                rule_time_str = log.RuleTime.isoformat() if log.RuleTime else ""
                ws.append([timestamp_str, log.StudentCode, log.Name, log.Subject, log.Section, log.Action, log.Status,
                           rule_time_str, log.Confidence, ""])
                ws.row_dimensions[row_num].height = 65
                if log.SnapshotPath and os.path.exists(log.SnapshotPath):
                    try:
                        pil_img = PILImage.open(log.SnapshotPath)
                        target_height = 80;
                        width_percent = (target_height / float(pil_img.size[1]));
                        target_width = int((float(pil_img.size[0]) * float(width_percent)));
                        pil_img = pil_img.resize((target_width, target_height), PILImage.LANCZOS)
                        img_io = io.BytesIO();
                        pil_img.save(img_io, format='PNG');
                        img_io.seek(0)
                        img_for_excel = OpenPyXLImage(img_io)
                        ws.add_image(img_for_excel, f"J{row_num}")
                    except Exception as e:
                        print(f"Error processing image {log.SnapshotPath}: {e}")
                        ws[f"J{row_num}"] = "Error: Img"
                else:
                    ws[f"J{row_num}"] = "N/A"
        wb.save(output)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:
        if not logs:
            df = pd.DataFrame(columns=headers_csv)
            df.loc[0] = ["No data found for the selected filters."] + [""] * 9
        else:
            log_dicts = [log._asdict() for log in logs]
            df = pd.DataFrame(log_dicts)
            df = df[headers_csv]
            if 'Timestamp' in df.columns:
                df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.tz_localize(None)
        df.to_csv(output, index=False, encoding='utf-8')
        media_type = "text/csv"
        filename += ".csv"
    return Response(
        content=output.getvalue(),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/attendance/clear/{cam_id}")
async def clear_attendance_log(
    cam_id: str,
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    if cam_mgr.clear_attendance_session(cam_id):
        return {"message": f"Attendance session for {cam_id} cleared."}
    else:
        raise HTTPException(status_code=404, detail=f"Camera {cam_id} not found or not open.")


# --- 8. User Management & Subject Endpoints ---
@app.get("/subjects", response_model=List[dict])
def list_subjects(
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    subjects = db.query(Subject).filter(Subject.is_deleted == 0).all()
    return [
        {"subject_id": s.subject_id, "subject_name": s.subject_name, "section": s.section,
         "cover_image_path": s.cover_image_path, "schedule": s.schedule,
         "academic_year": s.academic_year,
         "class_start_time": s.class_start_time.isoformat() if s.class_start_time else None,
         "class_end_time": s.class_end_time.isoformat() if s.class_end_time else None,
         }
        for s in subjects
    ]


class SubjectCreate(BaseModel):
    subject_name: str
    section: Optional[str] = None
    schedule: Optional[str] = None
    academic_year: Optional[str] = None
    class_start_time: Optional[dt_time] = None
    class_end_time: Optional[dt_time] = None


@app.post("/subjects", response_model=dict)
async def create_subject(
    payload: SubjectCreate,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    existing_subject = db.query(Subject).filter(
        Subject.subject_name == payload.subject_name,
        Subject.section == payload.section,
        Subject.academic_year == payload.academic_year
    ).first()
    if existing_subject:
        if existing_subject.is_deleted == 1:
            print(f"Undeleting subject: {payload.subject_name}")
            existing_subject.is_deleted = 0
            existing_subject.schedule = payload.schedule
            existing_subject.academic_year = payload.academic_year
            existing_subject.class_start_time = payload.class_start_time
            existing_subject.class_end_time = payload.class_end_time
            new_subject = existing_subject
        else:
            print(f"Subject already active: {payload.subject_name}")
            raise HTTPException(status_code=400, detail="Subject with this name/section/year already exists")
    else:
        print(f"Creating new subject: {payload.subject_name}")
        new_subject = Subject(
            subject_name=payload.subject_name,
            section=payload.section,
            schedule=payload.schedule,
            academic_year=payload.academic_year,
            class_start_time=payload.class_start_time,
            class_end_time=payload.class_end_time,
            is_deleted=0
        )
        db.add(new_subject)
    try:
        db.commit()
        db.refresh(new_subject)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {e}")
    return {
        "subject_id": new_subject.subject_id, "subject_name": new_subject.subject_name,
        "section": new_subject.section, "schedule": new_subject.schedule,
        "cover_image_path": new_subject.cover_image_path, "academic_year": new_subject.academic_year,
        "class_start_time": new_subject.class_start_time, "class_end_time": new_subject.class_end_time,
    }


@app.delete("/subjects/{subject_id}")
def delete_subject(
    subject_id: int,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    subject = db.query(Subject).filter(
        Subject.subject_id == subject_id,
        Subject.is_deleted == 0
    ).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    subject.is_deleted = 1
    db.commit()
    return {"message": f"Subject {subject_id} ({subject.subject_name}) marked as deleted."}


@app.get("/subjects/{subject_id}/student_count", response_model=dict)
def get_subject_student_count(
    subject_id: int,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    count = db.query(User).filter(
        User.subject_id == subject_id,
        User.is_deleted == 0
    ).count()
    return {"total_students": count}


class SubjectTimeUpdate(BaseModel):
    class_start_time: Optional[dt_time] = None


@app.put("/api/subjects/{subject_id}/time", response_model=dict)
def update_subject_time(
    subject_id: int,
    payload: SubjectTimeUpdate,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    subject = db.query(Subject).filter(Subject.subject_id == subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    try:
        subject.class_start_time = payload.class_start_time
        db.commit()
        print(f"Updated start time for subject {subject_id} to {payload.class_start_time}")
        return {"message": "Subject time updated successfully", "subject_id": subject_id,
                "new_time": payload.class_start_time}
    except Exception as e:
        db.rollback()
        print(f"Error updating subject time: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {e}")


class UserCreate(BaseModel):
    student_code: Optional[str] = None;
    name: str;
    role: str
    user_type_id: Optional[int] = None;
    subject_id: Optional[int] = None;
    password_hash: Optional[str] = None


class UserUpdate(BaseModel):
    name: Optional[str] = None;
    student_code: Optional[str] = None
    role: Optional[str] = None;
    subject_id: Optional[int] = None


@app.post("/users")
def create_user(
    payload: UserCreate,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    if payload.role not in ["admin", "operator", "viewer"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    if payload.student_code:
        existing = db.query(User).filter(User.student_code == payload.student_code, User.is_deleted == 0).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Student code '{payload.student_code}' already exists.")
    user = User(
        student_code=payload.student_code, name=payload.name, role=payload.role,
        user_type_id=payload.user_type_id, subject_id=payload.subject_id, password_hash=payload.password_hash,
    )
    db.add(user);
    db.commit();
    db.refresh(user)
    return {"message": "User created", "user": {
        "user_id": user.user_id, "student_code": user.student_code,
        "name": user.name, "role": user.role,
        "user_type_id": user.user_type_id,
        "subject_id": user.subject_id
    }}


@app.get("/users", response_model=List[dict])
def list_users(
    subject_id: Optional[int] = None,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    query = db.query(User).options(selectinload(User.faces)).filter(User.is_deleted == 0)
    if subject_id is not None:
        query = query.filter(User.subject_id == subject_id)
    users = query.all()
    results = []
    for u in users:
        results.append({
            "user_id": u.user_id, "student_code": u.student_code,
            "name": u.name, "role": u.role,
            "user_type_id": u.user_type_id,
            "subject_id": u.subject_id,
            "faces": [{"face_id": f.face_id, "file_path": f.file_path} for f in u.faces]
        })
    return results


@app.put("/users/{user_id}")
def update_user(
    user_id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    user = db.query(User).filter(User.user_id == user_id, User.is_deleted == 0).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    updated = False
    if payload.name is not None:
        user.name = payload.name;
        updated = True
    if payload.student_code is not None:
        if payload.student_code != user.student_code:
            existing = db.query(User).filter(User.student_code == payload.student_code, User.is_deleted == 0).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"Student code '{payload.student_code}' already exists.")
        user.student_code = payload.student_code;
        updated = True
    if payload.subject_id is not None:
        user.subject_id = payload.subject_id if payload.subject_id else None;
        updated = True
    if updated:
        db.commit()
    return {"message": "User updated", "user_id": user.user_id}


@app.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    user = db.query(User).options(selectinload(User.faces)).filter(User.user_id == user_id,
                                                                   User.is_deleted == 0).first()
    if not user: raise HTTPException(status_code=404, detail="User not found")
    user_face_dir = os.path.join(MEDIA_ROOT, str(user_id))
    try:
        if user.faces:
            for face in user.faces: db.delete(face)
        user.is_deleted = 1
        if user.student_code: user.student_code = f"{user.student_code}_deleted_{int(time.time())}"
        db.commit()
        if os.path.isdir(user_face_dir): shutil.rmtree(user_face_dir)
    except Exception as e:
        db.rollback();
        raise HTTPException(status_code=500, detail=f"Failed to delete user data: {e}")
    try:
        # ✨ [แก้ไข] เรียกใช้ Helper
        _internal_train_refresh(db)
    except Exception as e:
        print(f"Warning: _internal_train_refresh failed after deleting user: {e}")
    return {"message": f"User {user_id} ({user.name}) marked as deleted and all face data removed."}


@app.delete("/faces/{face_id}")
def delete_face(
    face_id: int,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    face = db.query(UserFace).filter(UserFace.face_id == face_id).first()
    if not face:
        raise HTTPException(status_code=404, detail="Face image not found")
    try:
        file_path = os.path.join(MEDIA_ROOT, str(face.user_id), face.file_path)
        db.delete(face);
        db.commit()
        if os.path.exists(file_path):
            os.remove(file_path)
        try:
            # ✨ [แก้ไข] เรียกใช้ Helper
            _internal_train_refresh(db)
        except Exception as e:
            print(f"Warning: _internal_train_refresh failed after deleting face: {e}")
        return {"message": f"Face image {face_id} ({face.file_path}) deleted."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete face: {e}")


# --- 9. Faculty Dashboard Endpoints ---
class ISubjectResponse(BaseModel): id: str; name: str
class ISemesterKPIs(BaseModel): totalRoster: int; avgAttendance: float; avgLateness: float; sessionsTaught: int
class ITrendDataset(BaseModel): label: str; data: List[float]; borderColor: Optional[str] = None; fill: Optional[
    bool] = False; backgroundColor: Optional[str] = None
class ITrendGraph(BaseModel): labels: List[str]; datasets: List[ITrendDataset]
class IStudentLateRisk(BaseModel):
    studentId: str
    name: str
    lates_percent: float
    lates_count: int
class IStudentAbsentRisk(BaseModel):
    studentId: str
    name: str
    absences_percent: float
    absences_count: int
class ISemesterOverviewData(BaseModel):
    kpis: ISemesterKPIs
    trendGraph: ITrendGraph
    studentsLate: List[IStudentLateRisk]
    studentsAbsent: List[IStudentAbsentRisk]
class ISessionKPIs(BaseModel): present: int; total: int; absent: int; late: int
class ISummaryDonutDataset(BaseModel): data: List[int]; backgroundColor: List[str]
class ISummaryDonut(BaseModel): labels: List[str]; datasets: List[ISummaryDonutDataset]
class IArrivalHistogramDataset(BaseModel): label: str; data: List[float]; backgroundColor: str
class IArrivalHistogram(BaseModel): labels: List[str]; datasets: List[IArrivalHistogramDataset]
class ILiveDataEntry(BaseModel): studentId: str; name: str; status: str; checkIn: Optional[str] = None; checkOut: \
    Optional[str] = None; duration: Optional[str] = None
class ISessionViewData(
    BaseModel): kpis: ISessionKPIs; summaryDonut: ISummaryDonut; arrivalHistogram: IArrivalHistogram; liveDataTable: \
    List[ILiveDataEntry]


@app.get("/api/faculty/subjects", response_model=List[ISubjectResponse])
def get_faculty_subjects(
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    print(f"User '{user_claims.get('name')}' requesting subjects...") # (ตัวอย่างการใช้ user_claims)
    try:
        subjects = db.query(Subject).filter(Subject.is_deleted == 0).order_by(Subject.academic_year.desc(),
                                                                              Subject.subject_name).all()
        results = []
        for s in subjects:
            name = f"[{s.academic_year or 'N/A'}] {s.subject_name}"
            if s.section: name += f" (Sec: {s.section})"
            results.append(ISubjectResponse(id=str(s.subject_id), name=name))
        if not results: raise HTTPException(status_code=404, detail="No subjects found. Please create a subject first.")
        return results
    except Exception as e:
        print(f"Database error in get_faculty_subjects: {e}")
        raise HTTPException(status_code=500,
                            detail=f"Database query failed. Check if DB schema is up to date. Error: {e}")


@app.get("/api/faculty/semester-overview", response_model=ISemesterOverviewData)
def get_semester_overview(
    subjectId: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    try:
        subject_id_int = int(subjectId)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid subjectId format.")

    roster_list = db.query(User).filter(User.subject_id == subject_id_int, User.is_deleted == 0).all()
    kpi_total_roster = len(roster_list)
    roster_map = {u.user_id: u for u in roster_list}
    if kpi_total_roster == 0:
        return ISemesterOverviewData(
            kpis=ISemesterKPIs(totalRoster=0, avgAttendance=0, avgLateness=0, sessionsTaught=0),
            trendGraph=ITrendGraph(labels=[], datasets=[]),
            studentsLate=[],
            studentsAbsent=[]
        )

    first_log_subquery = (db.query(
        AttendanceLog.user_id,
        func.date(AttendanceLog.timestamp).label("log_date"),
        func.min(AttendanceLog.timestamp).label("first_timestamp")
    ).filter(
        AttendanceLog.user_id.in_(roster_map.keys()),
        AttendanceLog.action == "enter"
    )
    )
    if start_date:
        first_log_subquery = first_log_subquery.filter(func.date(AttendanceLog.timestamp) >= start_date)
    if end_date:
        first_log_subquery = first_log_subquery.filter(func.date(AttendanceLog.timestamp) <= end_date)
    first_log_subquery = first_log_subquery.group_by(AttendanceLog.user_id,
                                                     func.date(AttendanceLog.timestamp)).subquery()

    all_first_logs = (db.query(
        first_log_subquery.c.user_id,
        first_log_subquery.c.log_date,
        AttendanceLog.log_status
    ).join(
        AttendanceLog,
        (AttendanceLog.user_id == first_log_subquery.c.user_id) &
        (AttendanceLog.timestamp == first_log_subquery.c.first_timestamp)
    )
                      .all()
                      )

    session_dates = sorted(list(set(log.log_date for log in all_first_logs)))
    kpi_sessions_taught = len(session_dates)
    if kpi_sessions_taught == 0:
        return ISemesterOverviewData(
            kpis=ISemesterKPIs(totalRoster=kpi_total_roster, avgAttendance=0, avgLateness=0, sessionsTaught=0),
            trendGraph=ITrendGraph(labels=[], datasets=[]),
            studentsLate=[],
            studentsAbsent=[]
        )

    student_lates_count = defaultdict(int)
    student_absences_count = defaultdict(int)
    logs_by_date = defaultdict(list)
    for user_id in roster_map.keys(): student_absences_count[user_id] = kpi_sessions_taught
    total_attendances = 0
    total_lates = 0

    for log in all_first_logs:
        logs_by_date[log.log_date].append(log)
        student_absences_count[log.user_id] -= 1
        total_attendances += 1
        if log.log_status == "Late":
            student_lates_count[log.user_id] += 1
            total_lates += 1

    total_possible_attendances = kpi_total_roster * kpi_sessions_taught
    kpi_avg_attendance = (total_attendances / total_possible_attendances) * 100 if total_possible_attendances > 0 else 0
    kpi_avg_lateness = (total_lates / total_attendances) * 100 if total_attendances > 0 else 0
    kpis = ISemesterKPIs(totalRoster=kpi_total_roster, avgAttendance=round(kpi_avg_attendance, 1),
                         avgLateness=round(kpi_avg_lateness, 1), sessionsTaught=kpi_sessions_taught)

    trend_dates_to_show = session_dates
    trend_labels = [d.strftime('%d/%m') for d in trend_dates_to_show]
    trend_data_present = [];
    trend_data_late = [];
    trend_data_absent = []

    for d in trend_dates_to_show:
        logs_on_date = logs_by_date[d]
        present_on_time_count = 0
        late_count = 0

        for log in logs_on_date:
            if log.log_status == "Late":
                late_count += 1
            elif log.log_status == "Present":
                present_on_time_count += 1

        attended_count = len(logs_on_date)
        absent_count = kpi_total_roster - attended_count

        present_percent = (present_on_time_count / kpi_total_roster) * 100
        late_percent = (late_count / kpi_total_roster) * 100
        absent_percent = (absent_count / kpi_total_roster) * 100

        trend_data_present.append(round(present_percent, 1))
        trend_data_late.append(round(late_percent, 1))
        trend_data_absent.append(round(absent_percent, 1))

    trend_graph = ITrendGraph(labels=trend_labels, datasets=[
        ITrendDataset(label="เข้าเรียน (%)", data=trend_data_present, borderColor='rgba(34, 197, 94, 1)',
                      backgroundColor='rgba(34, 197, 94, 0.1)', fill=True),
        ITrendDataset(label="สาย (%)", data=trend_data_late, borderColor='rgba(245, 158, 11, 1)',
                      backgroundColor='rgba(245, 158, 11, 0.1)', fill=True),
        ITrendDataset(label="ขาด (%)", data=trend_data_absent, borderColor='rgba(239, 68, 68, 1)',
                      backgroundColor='rgba(239, 68, 68, 0.1)', fill=True),
    ])

    students_late_list: List[IStudentLateRisk] = []
    students_absent_list: List[IStudentAbsentRisk] = []

    for user_id, user in roster_map.items():
        absences = student_absences_count[user_id]
        lates = student_lates_count[user_id]

        if absences > 0:
            absences_percent = (absences / kpi_sessions_taught) * 100 if kpi_sessions_taught > 0 else 0
            students_absent_list.append(
                IStudentAbsentRisk(
                    studentId=user.student_code or str(user.user_id),
                    name=user.name,
                    absences_percent=round(absences_percent, 1),
                    absences_count=absences
                )
            )

        if lates > 0:
            lates_percent = (lates / kpi_sessions_taught) * 100 if kpi_sessions_taught > 0 else 0
            students_late_list.append(
                IStudentLateRisk(
                    studentId=user.student_code or str(user.user_id),
                    name=user.name,
                    lates_percent=round(lates_percent, 1),
                    lates_count=lates
                )
            )

    students_absent_list.sort(key=lambda x: x.absences_percent, reverse=True)
    students_late_list.sort(key=lambda x: x.lates_percent, reverse=True)

    return ISemesterOverviewData(
        kpis=kpis,
        trendGraph=trend_graph,
        studentsLate=students_late_list,
        studentsAbsent=students_absent_list
    )


def add_minutes_to_time(t: dt_time, minutes_to_add: int) -> dt_time:
    dummy_date = date.today();
    full_dt = datetime.combine(dummy_date, t)
    new_dt = full_dt + timedelta(minutes=minutes_to_add)
    return new_dt.time()


@app.get("/api/faculty/session-view", response_model=ISessionViewData)
def get_session_view(
    subjectId: str,
    date: date,
    db: Session = Depends(get_db),
    # ✨ [ป้องกัน]
    user_claims: Dict[str, Any] = Depends(auth.get_token_claims)
):
    try:
        subject_id_int = int(subjectId)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid subjectId format.")

    subject = db.query(Subject).filter(Subject.subject_id == subject_id_int).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found.")
    CLASS_START_TIME = subject.class_start_time or dt_time(9, 0, 0)

    roster_list = db.query(User).filter(User.subject_id == subject_id_int, User.is_deleted == 0).all()
    kpi_total = len(roster_list);
    roster_map = {u.user_id: u for u in roster_list}

    def get_empty_histogram():
        empty_hist_labels = [f"{add_minutes_to_time(CLASS_START_TIME, (i - 2) * 5).strftime('%H:%M')}" for i in
                             range(7)]
        empty_hist_labels[2] = f"{CLASS_START_TIME.strftime('%H:%M')} (เริ่ม)"
        empty_hist_labels.insert(0, f"< {empty_hist_labels[0]}")
        return IArrivalHistogram(labels=empty_hist_labels, datasets=[
            IArrivalHistogramDataset(label="จำนวนนักเรียน", data=[0] * 8, backgroundColor='#6366f1')])

    if kpi_total == 0:
        print(f"Warning: No students enrolled in subject {subject_id_int}.")
        empty_kpis = ISessionKPIs(present=0, total=0, absent=0, late=0)
        empty_donut = ISummaryDonut(labels=["เข้าเรียน (0 คน, 0%)", "มาสาย (0 คน, 0%)", "ขาด (0 คน, 0%)"], datasets=[
            ISummaryDonutDataset(data=[0, 0, 0], backgroundColor=['#22c55e', '#f59e0b', '#ef4444'])])
        return ISessionViewData(kpis=empty_kpis, summaryDonut=empty_donut, arrivalHistogram=get_empty_histogram(),
                                liveDataTable=[])

    start_of_day = datetime.combine(date, dt_time.min);
    end_of_day = datetime.combine(date, dt_time.max)

    first_log_subquery = (
        db.query(AttendanceLog.user_id, func.min(AttendanceLog.timestamp).label("first_timestamp")).filter(
            AttendanceLog.user_id.in_(roster_map.keys()), AttendanceLog.action == "enter",
            AttendanceLog.timestamp.between(start_of_day, end_of_day)).group_by(AttendanceLog.user_id).subquery())
    first_logs_today = (db.query(
        AttendanceLog.user_id,
        AttendanceLog.timestamp,
        AttendanceLog.log_status
    ).join(first_log_subquery,
           (AttendanceLog.user_id == first_log_subquery.c.user_id) & (
                   AttendanceLog.timestamp == first_log_subquery.c.first_timestamp)).all())
    logs_map = {log.user_id: log for log in first_logs_today}

    live_data_table: List[ILiveDataEntry] = [];
    kpi_present_on_time = 0;
    kpi_late = 0;
    kpi_absent = 0

    for user_id, user in roster_map.items():
        log = logs_map.get(user_id)
        if log:
            status = log.log_status or "Present"
            if status == "Late":
                kpi_late += 1
            else:
                kpi_present_on_time += 1

            live_data_table.append(
                ILiveDataEntry(studentId=user.student_code or str(user.user_id), name=user.name, status=status,
                               checkIn=log.timestamp.isoformat(), checkOut=None, duration=None))
        else:
            kpi_absent += 1
            live_data_table.append(
                ILiveDataEntry(studentId=user.student_code or str(user.user_id), name=user.name, status="Absent",
                               checkIn=None, checkOut=None, duration=None))

    kpi_present_total = kpi_present_on_time + kpi_late
    kpis = ISessionKPIs(present=kpi_present_total, total=kpi_total, absent=kpi_absent, late=kpi_late)

    kpi_total_for_donut = kpi_total if kpi_total > 0 else 1
    p_present = (kpi_present_on_time / kpi_total_for_donut) * 100
    p_late = (kpi_late / kpi_total_for_donut) * 100
    p_absent = (kpi_absent / kpi_total_for_donut) * 100
    label_present = f"เข้าเรียน ({kpi_present_on_time} คน, {p_present:.0f}%)"
    label_late = f"มาสาย ({kpi_late} คน, {p_late:.0f}%)"
    label_absent = f"ขาด ({kpi_absent} คน, {p_absent:.0f}%)"
    summary_donut = ISummaryDonut(
        labels=[label_present, label_late, label_absent], datasets=[
            ISummaryDonutDataset(data=[kpi_present_on_time, kpi_late, kpi_absent],
                                 backgroundColor=['rgba(34, 197, 94, 0.7)', 'rgba(245, 158, 11, 0.7)',
                                                  'rgba(239, 68, 68, 0.7)'])])

    all_enter_logs_today = db.query(AttendanceLog.timestamp).filter(AttendanceLog.user_id.in_(roster_map.keys()),
                                                                    AttendanceLog.action == "enter",
                                                                    AttendanceLog.timestamp.between(start_of_day,
                                                                                                    end_of_day)).all()
    hist_labels = [];
    hist_buckets_start = [];
    interval_minutes = 5;
    num_buckets_before = 2;
    num_buckets_after = 4;
    total_buckets = num_buckets_before + 1 + num_buckets_after

    hist_labels_main = []
    hist_buckets_start_main = []
    for i in range(-num_buckets_before, num_buckets_after + 1):
        minutes = i * interval_minutes;
        bucket_time = add_minutes_to_time(CLASS_START_TIME, minutes)
        hist_buckets_start_main.append(bucket_time)
        if i == 0:
            hist_labels_main.append(f"{CLASS_START_TIME.strftime('%H:%M')} (เริ่ม)")
        else:
            hist_labels_main.append(bucket_time.strftime('%H:%M'))

    final_boundary = add_minutes_to_time(CLASS_START_TIME, (num_buckets_after + 1) * interval_minutes);
    hist_buckets_start_main.append(final_boundary);

    hist_data_count = [0] * (total_buckets + 1)
    hist_labels = [f"< {hist_labels_main[0]}"] + hist_labels_main

    main_start_time = hist_buckets_start_main[0]
    for log_tuple in all_enter_logs_today:
        log_time = log_tuple[0].time()

        if log_time < main_start_time:
            hist_data_count[0] += 1
        else:
            for i in range(total_buckets):
                if hist_buckets_start_main[i] <= log_time < hist_buckets_start_main[i + 1]:
                    hist_data_count[i + 1] += 1;
                    break

    arrival_histogram = IArrivalHistogram(labels=hist_labels, datasets=[
        IArrivalHistogramDataset(
            label="จำนวนนักเรียน",
            data=hist_data_count,
            backgroundColor='rgba(99, 102, 241, 0.7)')])

    return ISessionViewData(kpis=kpis, summaryDonut=summary_donut, arrivalHistogram=arrival_histogram,
                            liveDataTable=live_data_table)


# --- 10. Uvicorn Runner ---
if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)